from selenium import webdriver
from selenium.webdriver.common.by import By
from openpyxl import Workbook

# Khởi tạo trình duyệt Chrome
path_to_chromedriver = './chromedriver.exe'
driver = webdriver.Chrome(path_to_chromedriver)

# Mở trang web cần lấy thông tin
driver.get("https://howkteam.vn/learn")

lessons = driver.find_elements(By.CSS_SELECTOR, ".ribbon.ribbon-bookmark.ribbon-left.ribbon-success")

# for lesson in lessons:
#   authorList = lesson.find_elements(By.CSS_SELECTOR,".mr-10.useravatar-edit .ml-5")
#   for item in authorList:
#   	it = item.get_attribute('innerText');
#   	print (it);


#Tạo đối tượng Workbook
wb = Workbook()
dest_filename = 'crawl_data.xlsx';

# Lấy sheet đầu tiên
sheet = wb.active
sheet.title = 'Danh sách bài học'

# Ghi tiêu đề cho các cột
sheet["A1"] = "Tiêu đề"
sheet["B1"] = "Lượt xem"
sheet["C1"] = "Số bài học"
sheet["D1"] = "Tác giả"
sheet["E1"] = "Thumbnail"

# Duyệt qua danh sách các bài học và lấy thông tin
for i, lesson in enumerate(lessons):
  # Lấy thông tin từ các thành phần trên trang web
  title = lesson.find_element(By.CSS_SELECTOR,".block-content.block-content-full.block-sticky-options > h4").text
  view_count = lesson.find_element(By.CSS_SELECTOR,".si.si-eye.fa-fw ~ strong").text
  lesson_count = lesson.find_element(By.CSS_SELECTOR,".si.si-notebook.fa-fw ~ strong").text
  authorList = lesson.find_elements(By.CSS_SELECTOR,".block-content.block-content-full.useravatar-edit-container .mr-10.useravatar-edit .ml-5")
  author = '; '.join ([x.get_attribute('innerText') for x in authorList])
  # for item in authorList:
  # 	author += item.get_attribute('innerText') + '; '
  thumbnail = lesson.find_element(By.CSS_SELECTOR, ".img-fluid.options-item.w-100").get_attribute("src")


    # Ghi thông tin vào file Excel
  sheet.cell(row=i+2, column=1).value = title
  sheet.cell(row=i+2, column=2).value = view_count
  sheet.cell(row=i+2, column=3).value = lesson_count
  sheet.cell(row=i+2, column=4).value = author
  sheet.cell(row=i+2, column=5).value = thumbnail

  print ("Xong bài " + str(i) + "!");

wb.save(filename = dest_filename)